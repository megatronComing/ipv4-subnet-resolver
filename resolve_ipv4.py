'''
Resolve all subnets for a given IP address and save it to an excel file
'''
import ipaddress
import os
import sys
import openpyxl
from openpyxl.styles import Alignment

VERSION = '1.0'

def calculate_subnets(ip:str, subnet_mask_length:int) -> list[ipaddress.IPv4Network]:
    '''
    For a given subnet, obtain its next-level subnet (two subnets, that is, the subnet mask plus 1)
    对给定子网，获取其下一级子网(两个子网, 即子网掩码加1)
    Input:
        ip: str, an ipv4 address without subnet mask
        subnet_mask_length: int, the subnet mask length
    Return:
        a list of subnet objects
    '''
    network = ipaddress.IPv4Network(f'{ip}/{subnet_mask_length}', strict=False)
    subnets = list(network.subnets())
    return subnets

def to_binary(ip_address:str) -> str:
    '''
    Converts an decimal ipv4 address to binary
    Input:
        ip_address: str, an ipv4 address without subnet mask
    Return:
        dotted ipv4 address in binary
    
    '''
    binary = bin(int(ipaddress.IPv4Address(ip_address)))
    binary = binary[2:].zfill(32)
    return '.'.join([binary[i:i+8] for i in range(0, 32, 8)])

def subnet_details(subnet:ipaddress.IPv4Network) ->dict:
    '''
    Get detailed information about a given subnet
    获得指定子网的具体信息
    Input:
        subnet: a subnet object
    Return:
        a dictionary which contains information of the given address
    '''
    subnet_id = subnet.network_address
    subnet_id_binary = to_binary(subnet_id)
    
    hosts = list(subnet.hosts())
    first_host = hosts[0]
    first_host_binary = to_binary(first_host)

    last_host = hosts[-1]
    last_host_binary = to_binary(last_host)

    broadcast = subnet.broadcast_address
    broadcast_binary = to_binary(broadcast)

    subnet_mask = subnet.netmask
    subnet_mask_binary = to_binary(subnet_mask)

    return {
        "subnet_id": subnet_id,
        "subnet_id_bin": subnet_id_binary,
        "usable_hosts": len(hosts),
        "subnetmask": subnet_mask,
        "subnetmask_bin": subnet_mask_binary,
        "broadcast_addr": broadcast,
        "broadcast_addr_bin": broadcast_binary,
        "first_host": first_host,
        "first_host_bin": first_host_binary,
        "last_host": last_host,
        "last_host_bin": last_host_binary,
    }

def get_subnet(ip:str, subnetmask:int, indent:int, result:list):
    '''
    For a given subnet, obtain its next-level subnet (two subnets, that is, the subnet mask plus 1)
    对给定子网，获取其下一级子网(两个子网, 即子网掩码加1)
    Input:
        ip: str, an ipv4 address without subnet mask
        masklen: int, the subnet mask length
        indent: indentation level of the subnet id column of the excel file
        result: save the the subnets
    
    '''
    if subnetmask > 30:
        return
    seperator = ','
    indentation = ','  * indent
    #print(f'subnetting {ip}/{subnetmask}')
    subnets = calculate_subnets(ip, subnetmask)
    for subnet in subnets:
        # result['info'] = subnet_details(subnet)
        # subresult = {}
        # result['subnets'] = get_subnet(result['info']['subnet_id'], subnetmask+1, subresult)
        details = subnet_details(subnet)
        #print(f"{indentation}{details['subnet_id']}/{subnetmask+1}{seperator}{details['subnetmask']}{seperator}{details['usable_hosts']}{seperator}{details['first_host']}{seperator}{details['last_host']}")
        #result.append(f"{indentation}{details['subnet_id']}/{subnetmask+1}{seperator}{details['subnetmask']}{seperator}{details['usable_hosts']}{seperator}{details['first_host']}{seperator}{details['last_host']}")
        result.append({
            'indent': indent,
            'subnet_id': details['subnet_id'],
            'subnet_mask_len': subnetmask+1,
            'subnet_mask': details['subnetmask'],
            'usable_hosts': details['usable_hosts'],
            'first_host': details['first_host'],
            'last_host': details['last_host']
        })
        get_subnet(details['subnet_id'], subnetmask+1, indent + 1, result)

def subnet_and_save2file(ips:list, filename:str):
    '''
    Resolve all the subnets of the give IPv4 addresses and save it to an excel file.
    Input:
        ips: list of IPv4 addresses with subnet mask length, e.g. ["192.168.0.0/24", "201.100.20.0/24"]
        filename: excel file name to save the result into
    '''
    col_subn, col_subnm, col_hosts, col_firsth, col_lasth = 1,2,3,4,5
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=col_subn, value='Subnet ID')
    sheet.cell(row=1, column=col_subnm, value='Subnet Mask')
    sheet.cell(row=1, column=col_hosts, value='Usable Hosts')
    sheet.cell(row=1, column=col_firsth, value='First Host')
    sheet.cell(row=1, column=col_lasth, value='Last Host')
    lineno = 2
    
    for ip in ips:
        tmp = ip.split('/')
        ip_address = tmp[0]
        subnetlen = int(tmp[1])
        result = []
        get_subnet(ip_address, subnetlen, 1, result)
        sheet.cell(row=lineno, column=col_subn, value=ip)
        lineno += 1
        for line in result:
            sheet.cell(row=lineno, column=col_subn, value=f"{str(line['subnet_id'])}/{str(line['subnet_mask_len'])}")
            sheet.cell(row=lineno, column=col_subnm, value=str(line['subnet_mask']))
            sheet.cell(row=lineno, column=col_hosts, value=line['usable_hosts'])
            sheet.cell(row=lineno, column=col_firsth, value=str(line['first_host']))
            sheet.cell(row=lineno, column=col_lasth, value=str(line['last_host']))
            cell = sheet.cell(row=lineno, column=col_subn)
            cell.alignment = Alignment(indent=line['indent'] * 3)
            lineno += 1
    # save to the excel file
    try:
        workbook.save(filename)
        print(f'Save to file {filename}')
    except:
        print(f'FAILED to save to file {filename}')
def about() -> None:
    '''
    Shows the introduction info
    '''
    print('-'*20 + 'IPv4 aaddress subnetting ' + VERSION + '-'*20)
    print('any suggestion is welcome, contact hfyu.hzcn@gmail.com SVP.')
    print('-'*63)
def usage() -> None:
    '''
    Shows the usage info
    '''
    print('Usage: {} IP_ADDRESS1/SubnetMaskLength,IP_ADDRESS2/SubnetMaskLength [FILENAME]'.format(os.path.basename(sys.argv[0])))
    print('If OUTPUT is not given, will write to a default file in the current directory.')
def getArguments() -> dict:
    '''
    Parses the command line parameters
    Return:
        a dictionary
            item "ips": a list of ip addresses with subnet mask length. e.g. ["192.168.0.0/24", "201.100.20.0/24"]
            item "output": output file name
    '''
    argc = len(sys.argv)
    if argc == 1:
        return None
    paras = {}
    paras['ips'] = [ip.strip() for ip in sys.argv[1].split(',')]
    paras['output'] = sys.argv[2] if argc > 2 else "subnets.xlsx"

    return paras

if __name__=='__main__':
    about()
    args = getArguments()
    if args is None:
        usage()
        exit(-1)
    filename = args['output']
    #ips = ['196.10.10.0/24', '201.100.20.0/24']
    ips = args['ips']
    subnet_and_save2file(ips, filename)

